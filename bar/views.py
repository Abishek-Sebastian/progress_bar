from django.shortcuts import render
import openpyxl
from django.core.cache import cache
from .models import Person
from django.http import JsonResponse

def get_total_rows(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    return sheet.max_row

def insert_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    total_rows = sheet.max_row
    inserted_rows = 0

    for row in range(2, total_rows + 1):
        # Get the values from the Excel sheet
        name = sheet.cell(row, 1).value
        age = sheet.cell(row, 2).value
        city = sheet.cell(row, 3).value

        # Insert the data into the database
        person = Person(name=name, age=age, city=city)
        person.save()

        inserted_rows += 1

        # Update the progress indicator
        percentage = (inserted_rows / total_rows) * 100
        update_progress(percentage)

def update_progress(percentage):
    # Update the progress indicator here.
    # You could use Django's cache framework to store the percentage,
    # and then retrieve it in your template using JavaScript or jQuery.
    cache.set('import_progress', percentage, 60)

def import_data(request):
    file_path = request.FILES['file'].temporary_file_path()
    total_rows = get_total_rows(file_path)
    inserted_rows = insert_data(file_path)

    return render(request, 'import_data.html', {'total_rows': total_rows, 'inserted_rows': inserted_rows})

def get_progress(request):
    progress = cache.get('import_progress') or 0
    return JsonResponse({'progress': progress})

